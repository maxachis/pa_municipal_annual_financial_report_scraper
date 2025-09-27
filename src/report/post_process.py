from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font

from src.report.constants import OUTPUT_DIR
from src.report.enums import SheetName

import polars as pl

HIGHLIGHT_FILL = PatternFill(
    fgColor="FFFF00",
    fill_type="solid"
)
ZEBRA_FILL = PatternFill(
    fgColor="DDDDDD",  # Light gray
    fill_type="solid"
)

ROW_FONT = Font(
    name='Arial',
    size=11
)

def post_process(
    filename: str,
    county_muni_threshold: int,
):
    output_path: Path = OUTPUT_DIR / filename
    wb: Workbook = load_workbook(output_path)

    # Year and Muni


    ws_year_and_muni: Worksheet = wb[SheetName.YEAR_AND_MUNI.value]
    df_year_and_muni: pl.DataFrame = pl.read_excel(output_path, sheet_name=SheetName.YEAR_AND_MUNI.value)

    post_process_by_year(
        ws=ws_year_and_muni,
        df=df_year_and_muni,
        county_muni_threshold=county_muni_threshold
    )

    # Five-Year Avg
    if county_muni_threshold != 5:
        worksheet_name_five_year_avg: str = SheetName.FIVE_YEAR_AVG.value.replace("5", "4")
    else:
        worksheet_name_five_year_avg: str = SheetName.FIVE_YEAR_AVG.value

    ws_five_year_avg: Worksheet = wb[worksheet_name_five_year_avg]

    post_process_five_year_avg_and_per_capita(
        ws=ws_five_year_avg,
    )

    if county_muni_threshold != 5:
        worksheet_name_per_cap: str = SheetName.FIVE_YEAR_AVG.value.replace("5", "4")
    else:
        worksheet_name_per_cap: str = SheetName.FIVE_YEAR_AVG.value

    # Per Capita
    ws_per_capita: Worksheet = wb[worksheet_name_per_cap]
    post_process_five_year_avg_and_per_capita(ws_per_capita)

    wb.save(output_path)



def post_process_by_year(
    ws: Worksheet,
    df: pl.DataFrame,
    county_muni_threshold: int,
) -> None:

    # Get all county/muni groups with fewer than county_muni_threshold rows
    df_county_muni = (
        df
        .group_by(["County", "Municipality"])
        .agg(pl.count())
        .filter(pl.col("count") < county_muni_threshold)
    )
    # Create a set of county/muni pairs
    county_muni_set = set(
        zip(
            df_county_muni["County"],
            df_county_muni["Municipality"]
        )
    )

    # Format rows with those values with a highlight
    # Iterate through worksheet rows
    for row in ws.iter_rows(min_row=2):
        # Set font to Arial 11
        for cell in row:
            cell.font = ROW_FONT

        # Conditionally highlight rows if county/muni is in the set
        county = row[0].value
        municipality = row[1].value
        if (county, municipality) in county_muni_set:
            for cell in row:
                cell.fill = HIGHLIGHT_FILL

def post_process_five_year_avg_and_per_capita(
    ws: Worksheet,
) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            # Optionally apply zebra striping
            if cell.row % 2 == 0:
                cell.fill = ZEBRA_FILL

            # Set font to Arial 11
            cell.font = ROW_FONT

