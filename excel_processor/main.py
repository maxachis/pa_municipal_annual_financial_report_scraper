"""
This page contains both the ExcelProcessor class
and the main function for processing excel files

"""

import re
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.exc import IntegrityError

from config import REPORT_RELEVANT_SHEET_NAME, REL_TOTAL_COLUMN, REL_CODE_COLUMN, REL_LABEL_COLUMN, \
    JOINED_POP_RELEVANT_SHEET_NAME, JOINED_GEO_COLUMN, JOINED_MUNI_COLUMN, JOINED_COUNTY_COLUMN, JOINED_CLASS_COLUMN, \
    JOINED_POP_ESTIMATE_COLUMN, JOINED_URBAN_RURAL_COLUMN, JOINED_POP_MARGIN_COLUMN
from database_logic.DatabaseManager import DatabaseManager
from database_logic.models import JoinedPopDetails
from scraper.JsonCache import JsonCache
from scraper.data_objects import CMY
from util import project_path

MAX_ROW = 1000
VALID_ROW_REGEX = r"^[\d\-\.]+$"  # Regex for row containing only numbers, decimals, and hyphens

def open_excel_file(file_path: Path) -> Workbook:
    return load_workbook(file_path)

def case_insensitive_replace(text, old, new):
    pattern = re.compile(re.escape(old), re.IGNORECASE)
    return pattern.sub(new, text)

class ExcelProcessor:
    """
    Class for processing scraped excel files and adding them to the database
    """

    def __init__(self):
        self.database_manager = DatabaseManager()
        self.cache = JsonCache()
        self.cache.load_cache()

    def process_downloaded_reports(self):
        for cmy in self.cache.get_as_list_of_CMY():
            print(f"Processing {cmy.county} {cmy.municipality} {cmy.year}")
            try:
                wb = self.get_downloaded_report(cmy)
                self.process_downloaded_report(wb, cmy)
                self.cache.mark_as_processed(cmy)
            except Exception as e:
                print(f"Error processing {cmy.county} {cmy.municipality} {cmy.year}: {e}")
                self.cache.add_process_error(cmy, str(e))

    def get_downloaded_report(self, cmy: CMY) -> Workbook:
        return open_excel_file(
            file_path=project_path(
                "downloads",
                f"report_{cmy.county}_{cmy.municipality}_{cmy.year}.xlsx"
            )
        )

    def process_downloaded_report(self, wb: Workbook, cmy: CMY):
        # Get sheet
        ws: Worksheet = wb[REPORT_RELEVANT_SHEET_NAME]

        # Iterate through rows until getting to end
        for row in ws.iter_rows(
                min_row=1,
                max_row=MAX_ROW,
                min_col=1,
                max_col=REL_TOTAL_COLUMN,
                values_only=True
        ):
            if row[0] is None:
                continue
            if not re.match(VALID_ROW_REGEX, row[0]):
                continue
            code = row[REL_CODE_COLUMN - 1]
            label = row[REL_LABEL_COLUMN - 1]
            total = row[REL_TOTAL_COLUMN - 1]
            if not self.database_manager.code_label_exists(code):
                self.database_manager.add_code_label(code, label)
            try:
                self.database_manager.add_to_annual_financial_report_details_table(
                    county=cmy.county,
                    municipality=cmy.municipality,
                    year=cmy.year,
                    code=code,
                    total=total
                )
            except IntegrityError:
                pass

    def clean_county(self, county: str) -> str:
        return case_insensitive_replace(county, "County", "")

    def clean_municipality(self, municipality: str) -> str:
        for term in ("Borough", "City", "Township"):
            return case_insensitive_replace(municipality, term, "")

    def process_joined_pop_class_urban_rural(self):
        self.database_manager.wipe_table(JoinedPopDetails)
        wb = open_excel_file("Joined pop class urban rural.xlsx")


        ws: Worksheet = wb[JOINED_POP_RELEVANT_SHEET_NAME]
        for row in ws.iter_rows(min_row=2, max_row=MAX_ROW, min_col=1, max_col=8, values_only=True):
            if row[0] is None:
                continue
            if row[0].strip() == "":
                continue
            geo = row[JOINED_GEO_COLUMN - 1]
            muni = row[JOINED_MUNI_COLUMN - 1]
            muni = self.clean_municipality(muni)
            county = row[JOINED_COUNTY_COLUMN - 1]
            county = self.clean_county(county)
            class_ = row[JOINED_CLASS_COLUMN - 1]
            pop_estimate = row[JOINED_POP_ESTIMATE_COLUMN - 1]
            pop_margin = row[JOINED_POP_MARGIN_COLUMN - 1]
            rural_urban = row[JOINED_URBAN_RURAL_COLUMN - 1]
            try:
                self.database_manager.add_pop_row(
                    geo_id=geo,
                    county=county,
                    municipality=muni,
                    class_=class_,
                    pop_estimate=pop_estimate,
                    pop_margin=pop_margin,
                    urban_rural=rural_urban
                )
            except IntegrityError:
                pass





if __name__ == "__main__":
    processor = ExcelProcessor()
    processor.process_downloaded_reports()
    processor.process_joined_pop_class_urban_rural()